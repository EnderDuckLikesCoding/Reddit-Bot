import praw
import openpyxl as xl
import time

index = 0
delay = 3
filename = input('filename (*.xslx): ')
sheetname = input('Sheet Name: ')
subject = input('Subject: ')
message = input('Message: ')


wb = xl.load_workbook(filename)
sheet = wb[sheetname]


names = []
for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=1):
    for cell in row:
        if cell.value is not None:
            names.append(cell.value)


reddit = praw.Reddit(client_id='-_jp3B32nE2jxw',
                     client_secret='qfo3JrQGI0vcm7s4W2u9qrsQDoI',
                     username='BotTesting11',
                     password='BotPassword:123',
                     user_agent='user agent'
                     )
while True:
    index += 1;
    time.sleep(delay)
    reddit.redditor(names[index]).message(subject, message)
    if index >= len(names) - 1:
        break